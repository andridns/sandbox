"""
Import API endpoints for Excel file imports
"""
import logging
import re
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from typing import List, Dict
from pathlib import Path
from app.database import get_db
from app.models.expense import Expense
from app.models.category import Category
from app.models.user import User
from app.services.excel_import import ExcelImportService
from app.services.category_matcher import CategoryMatcher
from app.schemas.expense import ExpenseCreate
from app.schemas.category import CategoryCreate
from app.core.auth import get_current_user
from decimal import Decimal
from datetime import date, datetime
from openpyxl import load_workbook
import io
from uuid import UUID

logger = logging.getLogger(__name__)

router = APIRouter()

# Max file size: 10MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Allowed extensions
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Import expenses from Excel file with smart categorization
    """
    try:
        if not file.filename:
            raise HTTPException(
                status_code=400,
                detail="Filename is required"
            )
        
        logger.info(f"Starting Excel import. Filename: {file.filename}")
        
        # Check file extension
        file_ext = Path(file.filename).suffix.lower()
        logger.debug(f"File extension: {file_ext}")
        if file_ext not in ALLOWED_EXTENSIONS:
            error_msg = f"Invalid file type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
            logger.error(error_msg)
            raise HTTPException(
                status_code=400,
                detail=error_msg
            )
        
        # Read file content
        logger.debug("Reading file content")
        contents = await file.read()
        file_size = len(contents)
        logger.info(f"File read successfully. Size: {file_size} bytes ({file_size / 1024:.2f} KB)")
        
        # Check file size
        if file_size > MAX_FILE_SIZE:
            error_msg = f"File too large. Maximum size: {MAX_FILE_SIZE / 1024 / 1024}MB"
            logger.error(f"{error_msg}. Actual size: {file_size / 1024 / 1024:.2f}MB")
            raise HTTPException(
                status_code=400,
                detail=error_msg
            )
        
        # Load workbook to check for Categories sheet
        logger.info("Loading workbook to check for Categories sheet")
        workbook = load_workbook(io.BytesIO(contents), data_only=True)
        
        # Import categories if Categories sheet exists
        categories_imported = 0
        category_id_map = {}  # Map old IDs to new IDs
        if "Categories" in workbook.sheetnames:
            logger.info("Found Categories sheet, importing categories")
            ws_categories = workbook["Categories"]
            
            # Find header row
            header_row = None
            for row_idx in range(1, min(11, ws_categories.max_row + 1)):
                row = ws_categories[row_idx]
                headers = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
                if "name" in headers or "id" in headers:
                    header_row = row_idx
                    break
            
            if header_row:
                # Map column indices
                col_map = {}
                header_cells = ws_categories[header_row]
                for idx, cell in enumerate(header_cells, start=1):
                    if cell.value:
                        header_name = str(cell.value).strip().lower()
                        col_map[header_name] = idx
                
                # Process category rows
                for row_idx in range(header_row + 1, ws_categories.max_row + 1):
                    row = ws_categories[row_idx]
                    if not any(cell.value for cell in row):
                        continue
                    
                    try:
                        old_id = None
                        name = None
                        icon = None
                        color = "#4CAF50"
                        is_default = False
                        
                        if "id" in col_map:
                            id_cell = row[col_map["id"] - 1]
                            if id_cell.value:
                                old_id = str(id_cell.value).strip()
                        
                        if "name" in col_map:
                            name_cell = row[col_map["name"] - 1]
                            if name_cell.value:
                                name = str(name_cell.value).strip()
                        
                        if "icon" in col_map:
                            icon_cell = row[col_map["icon"] - 1]
                            if icon_cell.value:
                                icon = str(icon_cell.value).strip() or None
                        
                        if "color" in col_map:
                            color_cell = row[col_map["color"] - 1]
                            if color_cell.value:
                                color = str(color_cell.value).strip()
                        
                        if "is default" in col_map:
                            default_cell = row[col_map["is default"] - 1]
                            if default_cell.value:
                                is_default = str(default_cell.value).strip().lower() in ["yes", "true", "1"]
                        
                        if name:
                            # Check if category already exists
                            existing = db.query(Category).filter(Category.name == name).first()
                            if existing:
                                if old_id:
                                    category_id_map[old_id] = existing.id
                                logger.debug(f"Category '{name}' already exists, skipping")
                            else:
                                # Create new category
                                new_category = Category(
                                    name=name,
                                    icon=icon,
                                    color=color,
                                    is_default=is_default
                                )
                                db.add(new_category)
                                db.commit()
                                db.refresh(new_category)
                                
                                if old_id:
                                    category_id_map[old_id] = new_category.id
                                
                                categories_imported += 1
                                logger.info(f"Imported category: {name}")
                    except Exception as e:
                        logger.warning(f"Failed to import category from row {row_idx}: {e}")
                        continue
        
        # Parse Excel file for expenses
        logger.info("Parsing Excel file for expenses")
        import_service = ExcelImportService(contents)
        expenses_data, parse_errors = import_service.parse()
        logger.info(f"Excel parsing complete. Found {len(expenses_data)} valid rows, {len(parse_errors)} parse errors")
        
        if parse_errors:
            logger.warning(f"Parse errors encountered: {parse_errors[:5]}")  # Log first 5 errors
        
        if not expenses_data and not parse_errors and categories_imported == 0:
            error_msg = "No data found in Excel file"
            logger.error(error_msg)
            raise HTTPException(
                status_code=400,
                detail=error_msg
            )
        
        # Initialize category matcher
        logger.debug("Initializing category matcher")
        category_matcher = CategoryMatcher(db)
        
        # Process expenses
        imported_count = 0
        failed_rows: List[Dict] = []
        category_matches: Dict[str, int] = {}
        uncategorized_count = 0
        
        # Get all categories for direct name matching
        logger.debug("Loading categories from database")
        all_categories = db.query(Category).all()
        # Create maps for matching: exact name, lowercase name, and name without emojis
        category_name_map = {}
        category_name_lower_map = {cat.name.lower(): cat for cat in all_categories}
        logger.info(f"Loaded {len(all_categories)} categories from database")
        
        # Helper function to strip emojis and normalize category name
        def normalize_category_name(category_str: str) -> str:
            """Strip emojis and normalize category name for matching"""
            # Remove emojis (Unicode emoji ranges)
            emoji_pattern = re.compile(
                "["
                "\U0001F600-\U0001F64F"  # emoticons
                "\U0001F300-\U0001F5FF"  # symbols & pictographs
                "\U0001F680-\U0001F6FF"  # transport & map symbols
                "\U0001F1E0-\U0001F1FF"  # flags (iOS)
                "\U00002702-\U000027B0"
                "\U000024C2-\U0001F251"
                "]+", flags=re.UNICODE
            )
            normalized = emoji_pattern.sub('', category_str).strip()
            return normalized
        
        skipped_count = 0
        
        logger.info(f"Processing {len(expenses_data)} expense rows")
        for idx, expense_data in enumerate(expenses_data, start=1):
            try:
                # Skip if amount is 0 or negative (like income entries, taxes, etc.)
                amount = expense_data.get('amount', 0)
                if amount <= 0:
                    skipped_count += 1
                    logger.debug(f"Row {idx + 1}: Skipping row with amount {amount}")
                    continue
                
                logger.debug(f"Row {idx + 1}: Processing expense - Amount: {amount}, Description: {expense_data.get('description', 'N/A')[:50]}")
                matched_category = None
                
                # Check if expense has a category ID from the export (if ID column exists)
                expense_category_id = None
                if 'id' in expense_data and expense_data.get('id'):
                    # Try to find expense by ID to get its category_id
                    try:
                        expense_id = UUID(str(expense_data['id']))
                        existing_expense = db.query(Expense).filter(Expense.id == expense_id).first()
                        if existing_expense and existing_expense.category_id:
                            expense_category_id = existing_expense.category_id
                            logger.debug(f"Row {idx + 1}: Found existing expense with category_id: {expense_category_id}")
                    except (ValueError, TypeError):
                        pass
                
                # Match category directly from Excel file (categories are already in correct format)
                if 'category' in expense_data and expense_data['category']:
                    category_value = str(expense_data['category']).strip()
                    logger.debug(f"Row {idx + 1}: Category value from file: '{category_value}'")
                    
                    # Try exact match first (case-insensitive)
                    category_value_lower = category_value.lower()
                    if category_value_lower in category_name_lower_map:
                        matched_category = category_name_lower_map[category_value_lower]
                        logger.info(f"Row {idx + 1}: Matched category via exact name: {matched_category.name}")
                    else:
                        # Strip emojis and try matching again
                        normalized_category = normalize_category_name(category_value)
                        normalized_lower = normalized_category.lower()
                        logger.debug(f"Row {idx + 1}: Normalized category (no emojis): '{normalized_category}'")
                        
                        # Try exact match with normalized name
                        if normalized_lower in category_name_lower_map:
                            matched_category = category_name_lower_map[normalized_lower]
                            logger.info(f"Row {idx + 1}: Matched category via normalized name: {matched_category.name}")
                        else:
                            # Try partial/fuzzy match (check if normalized category contains or is contained in any category name)
                            for cat_name_lower, category in category_name_lower_map.items():
                                # Check if category name contains the normalized value or vice versa
                                if normalized_lower in cat_name_lower or cat_name_lower in normalized_lower:
                                    matched_category = category
                                    logger.info(f"Row {idx + 1}: Matched category via fuzzy match: '{normalized_category}' -> '{matched_category.name}'")
                                    break
                
                # If we found a category_id from existing expense, use it (mapped if needed)
                if expense_category_id and not matched_category:
                    # Check if this ID was mapped during category import
                    mapped_id = category_id_map.get(str(expense_category_id))
                    if mapped_id:
                        matched_category = db.query(Category).filter(Category.id == mapped_id).first()
                        if matched_category:
                            logger.info(f"Row {idx + 1}: Using mapped category from import: {matched_category.name}")
                    else:
                        # Try to find category by original ID
                        matched_category = db.query(Category).filter(Category.id == expense_category_id).first()
                        if matched_category:
                            logger.info(f"Row {idx + 1}: Using category from expense ID: {matched_category.name}")
                
                if matched_category:
                    expense_data['category_id'] = matched_category.id
                    category_name = matched_category.name
                    category_matches[category_name] = category_matches.get(category_name, 0) + 1
                    logger.debug(f"Row {idx + 1}: Category assigned: {category_name} (ID: {matched_category.id})")
                else:
                    expense_data['category_id'] = None
                    uncategorized_count += 1
                    logger.debug(f"Row {idx + 1}: No category matched, leaving uncategorized")
                
                # Ensure date is a date object (not datetime) before database import
                # Extract exact date from source datetime if needed
                expense_date = expense_data['date']
                if isinstance(expense_date, datetime):
                    expense_date = expense_date.date()
                    logger.debug(f"Row {idx + 1}: Converted datetime to date: {expense_date}")
                elif not isinstance(expense_date, date):
                    # Fallback to today if somehow not a date/datetime
                    expense_date = date.today()
                    logger.warning(f"Row {idx + 1}: Invalid date type, using today: {expense_date}")
                
                # Convert to ExpenseCreate format
                expense_create = ExpenseCreate(
                    amount=expense_data['amount'],
                    currency=expense_data.get('currency', 'IDR'),
                    description=expense_data['description'],
                    category_id=expense_data.get('category_id'),
                    date=expense_date,
                    tags=expense_data.get('tags', []),
                    location=expense_data.get('location'),
                    notes=expense_data.get('notes'),
                    is_recurring=False
                )
                logger.debug(f"Row {idx + 1}: Created ExpenseCreate object with date: {expense_date}")
                
                # Create expense
                logger.debug(f"Row {idx + 1}: Saving to database")
                db_expense = Expense(**expense_create.model_dump())
                db.add(db_expense)
                db.commit()
                
                imported_count += 1
                logger.info(f"Row {idx + 1}: Successfully imported expense (ID: {db_expense.id})")
                
            except Exception as e:
                db.rollback()
                error_info = {
                    "row": idx + 1,  # +1 because we start from row 2 (row 1 is header)
                    "error": str(e),
                    "data": expense_data
                }
                logger.exception(f"Row {idx + 1}: Failed to import expense - {str(e)}")
                logger.debug(f"Row {idx + 1}: Failed row data: {expense_data}")
                failed_rows.append(error_info)
        
        # Prepare response
        summary = {
            "total_rows": len(expenses_data),
            "imported": imported_count,
            "failed": len(failed_rows) + len(parse_errors),
            "uncategorized": uncategorized_count,
            "skipped": skipped_count,
            "categories_imported": categories_imported,
        }
        
        logger.info(f"Import complete. Summary: {summary}")
        logger.info(f"Category matches: {category_matches}")
        if failed_rows:
            logger.warning(f"Failed rows: {len(failed_rows)}")
            for failed_row in failed_rows[:5]:  # Log first 5 failed rows
                logger.warning(f"Failed row {failed_row['row']}: {failed_row['error']}")
        
        return {
            "success": True,
            "summary": summary,
            "category_matches": category_matches,
            "errors": parse_errors + [f"Row {row['row']}: {row['error']}" for row in failed_rows],
            "failed_rows": failed_rows[:10]  # Limit to first 10 failed rows
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error during Excel import: {str(e)}")
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Full traceback: {error_details}")
        raise HTTPException(
            status_code=500,
            detail=f"Error importing Excel file: {str(e)}"
        )
