from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.database import get_db
from app.models.expense import Expense
from app.models.category import Category
from app.models.user import User
from app.core.auth import get_current_user

router = APIRouter()


@router.get("/export/csv")
async def export_csv(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export expenses to CSV"""
    query = db.query(Expense)
    
    if start_date:
        query = query.filter(Expense.date >= start_date)
    if end_date:
        query = query.filter(Expense.date <= end_date)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Date", "Amount", "Currency", "Description", "Category", 
        "Tags", "Location", "Recurring", "Notes"
    ])
    
    # Data rows
    for expense in expenses:
        writer.writerow([
            expense.date.isoformat(),
            str(expense.amount),
            expense.currency,
            expense.description,
            expense.category.name if expense.category else "",
            ", ".join(expense.tags) if expense.tags else "",
            expense.location or "",
            "Yes" if expense.is_recurring else "No",
            expense.notes or ""
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=expenses.csv"}
    )


@router.get("/export/excel")
async def export_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export expenses to Excel"""
    query = db.query(Expense)
    
    if start_date:
        query = query.filter(Expense.date >= start_date)
    if end_date:
        query = query.filter(Expense.date <= end_date)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create Expenses sheet
    ws_expenses = wb.create_sheet("Expenses")
    
    # Expenses header
    expense_headers = [
        "ID", "Date", "Amount", "Currency", "Description", "Category",
        "Tags", "Location", "Recurring", "Notes"
    ]
    ws_expenses.append(expense_headers)
    
    # Style expenses header
    for cell in ws_expenses[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Expenses data rows
    for expense in expenses:
        ws_expenses.append([
            str(expense.id),
            expense.date.isoformat(),
            float(expense.amount),
            expense.currency,
            expense.description,
            expense.category.name if expense.category else "",
            ", ".join(expense.tags) if expense.tags else "",
            expense.location or "",
            "Yes" if expense.is_recurring else "No",
            expense.notes or ""
        ])
    
    # Auto-adjust expenses column widths
    for column in ws_expenses.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_expenses.column_dimensions[column_letter].width = adjusted_width
    
    # Create Categories sheet
    categories = db.query(Category).order_by(Category.is_default.desc(), Category.name).all()
    ws_categories = wb.create_sheet("Categories")
    
    # Categories header
    category_headers = [
        "ID", "Name", "Icon", "Color", "Is Default"
    ]
    ws_categories.append(category_headers)
    
    # Style categories header
    for cell in ws_categories[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Categories data rows
    for category in categories:
        ws_categories.append([
            str(category.id),
            category.name,
            category.icon or "",
            category.color,
            "Yes" if category.is_default else "No"
        ])
    
    # Auto-adjust categories column widths
    for column in ws_categories.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_categories.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=expenses.xlsx"}
    )


@router.get("/export/pdf")
async def export_pdf(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export expenses to PDF"""
    query = db.query(Expense)
    
    if start_date:
        query = query.filter(Expense.date >= start_date)
    if end_date:
        query = query.filter(Expense.date <= end_date)
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("Expense Report", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Date range
    if start_date or end_date:
        date_range = f"From {start_date or 'Beginning'} to {end_date or 'Today'}"
        story.append(Paragraph(date_range, styles["Normal"]))
        story.append(Spacer(1, 12))
    
    # Table data
    data = [["Date", "Amount", "Description", "Category"]]
    
    for expense in expenses:
        data.append([
            expense.date.strftime("%Y-%m-%d"),
            f"{expense.currency} {expense.amount:,.2f}",
            expense.description[:30] + "..." if len(expense.description) > 30 else expense.description,
            expense.category.name if expense.category else "N/A"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
    ]))
    
    story.append(table)
    
    doc.build(story)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=expenses.pdf"}
    )


@router.get("/export/count")
async def get_expense_count(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get the total count of expenses/transactions in the database"""
    count = db.query(func.count(Expense.id)).scalar()
    return {"count": count}
