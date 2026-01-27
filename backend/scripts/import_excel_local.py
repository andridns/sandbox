"""
Local Excel Import Script for Production Database

This script allows importing Excel files directly to production database from your local machine.
It processes expenses in batches to avoid timeout issues.

Usage:
    export DATABASE_URL="postgresql://user:pass@host:port/dbname"
    poetry run python scripts/import_excel_local.py /path/to/file.xlsx

Or with options:
    poetry run python scripts/import_excel_local.py file.xlsx --batch-size 500 --skip-existing
"""
import sys
import argparse
import os
import logging
import re
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID
from openpyxl import load_workbook
import io

# Add parent directory to path to import app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.database import Base
from app.models.expense import Expense
from app.models.category import Category
from app.services.excel_import import ExcelImportService
from app.schemas.expense import ExpenseCreate

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def normalize_category_name(category_str: str) -> str:
    """Strip emojis and normalize category name for matching"""
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags (iOS)
        "\U00002702-\U000027B0"
        "\U000024C2-\U0001F251"
        "]+", flags=re.UNICODE
    )
    normalized = emoji_pattern.sub('', category_str).strip()
    return normalized


def import_categories(workbook, db, category_id_map: Dict[str, UUID]) -> int:
    """Import categories from Categories sheet if it exists"""
    categories_imported = 0
    
    if "Categories" not in workbook.sheetnames:
        logger.info("No Categories sheet found, skipping category import")
        return categories_imported
    
    logger.info("Found Categories sheet, importing categories")
    ws_categories = workbook["Categories"]
    
    # Find header row
    header_row = None
    for row_idx in range(1, min(11, ws_categories.max_row + 1)):
        row = ws_categories[row_idx]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in row]
        if "name" in headers or "id" in headers:
            header_row = row_idx
            break
    
    if not header_row:
        logger.warning("Could not find header row in Categories sheet")
        return categories_imported
    
    # Map column indices
    col_map = {}
    header_cells = ws_categories[header_row]
    for idx, cell in enumerate(header_cells, start=1):
        if cell.value:
            header_name = str(cell.value).strip().lower()
            col_map[header_name] = idx
    
    # Process category rows
    for row_idx in range(header_row + 1, ws_categories.max_row + 1):
        row = ws_categories[row_idx]
        if not any(cell.value for cell in row):
            continue
        
        try:
            old_id = None
            name = None
            icon = None
            color = "#4CAF50"
            is_default = False
            
            if "id" in col_map:
                id_cell = row[col_map["id"] - 1]
                if id_cell.value:
                    old_id = str(id_cell.value).strip()
            
            if "name" in col_map:
                name_cell = row[col_map["name"] - 1]
                if name_cell.value:
                    name = str(name_cell.value).strip()
            
            if "icon" in col_map:
                icon_cell = row[col_map["icon"] - 1]
                if icon_cell.value:
                    icon = str(icon_cell.value).strip() or None
            
            if "color" in col_map:
                color_cell = row[col_map["color"] - 1]
                if color_cell.value:
                    color = str(color_cell.value).strip()
            
            if "is default" in col_map:
                default_cell = row[col_map["is default"] - 1]
                if default_cell.value:
                    is_default = str(default_cell.value).strip().lower() in ["yes", "true", "1"]
            
            if name:
                # Check if category already exists
                existing = db.query(Category).filter(Category.name == name).first()
                if existing:
                    if old_id:
                        category_id_map[old_id] = existing.id
                    logger.debug(f"Category '{name}' already exists, skipping")
                else:
                    # Create new category
                    new_category = Category(
                        name=name,
                        icon=icon,
                        color=color,
                        is_default=is_default
                    )
                    db.add(new_category)
                    db.commit()
                    db.refresh(new_category)
                    
                    if old_id:
                        category_id_map[old_id] = new_category.id
                    
                    categories_imported += 1
                    logger.info(f"Imported category: {name}")
        except Exception as e:
            logger.warning(f"Failed to import category from row {row_idx}: {e}")
            db.rollback()
            continue
    
    return categories_imported


def match_category(expense_data: Dict, all_categories: List[Category], 
                   category_id_map: Dict[str, UUID], db) -> Optional[Category]:
    """Match category for an expense using the same logic as production API"""
    matched_category = None
    
    # Create lowercase category name map
    category_name_lower_map = {cat.name.lower(): cat for cat in all_categories}
    
    # Check if expense has a category ID from the export
    expense_category_id = None
    if 'id' in expense_data and expense_data.get('id'):
        try:
            expense_id = UUID(str(expense_data['id']))
            existing_expense = db.query(Expense).filter(Expense.id == expense_id).first()
            if existing_expense and existing_expense.category_id:
                expense_category_id = existing_expense.category_id
        except (ValueError, TypeError):
            pass
    
    # Match category directly from Excel file
    if 'category' in expense_data and expense_data['category']:
        category_value = str(expense_data['category']).strip()
        
        # Try exact match first (case-insensitive)
        category_value_lower = category_value.lower()
        if category_value_lower in category_name_lower_map:
            matched_category = category_name_lower_map[category_value_lower]
        else:
            # Strip emojis and try matching again
            normalized_category = normalize_category_name(category_value)
            normalized_lower = normalized_category.lower()
            
            # Try exact match with normalized name
            if normalized_lower in category_name_lower_map:
                matched_category = category_name_lower_map[normalized_lower]
            else:
                # Try partial/fuzzy match
                for cat_name_lower, category in category_name_lower_map.items():
                    if normalized_lower in cat_name_lower or cat_name_lower in normalized_lower:
                        matched_category = category
                        break
    
    # If we found a category_id from existing expense, use it (mapped if needed)
    if expense_category_id and not matched_category:
        mapped_id = category_id_map.get(str(expense_category_id))
        if mapped_id:
            matched_category = db.query(Category).filter(Category.id == mapped_id).first()
        else:
            matched_category = db.query(Category).filter(Category.id == expense_category_id).first()
    
    return matched_category


def process_batch(expenses_batch: List[Dict], all_categories: List[Category],
                  category_id_map: Dict[str, UUID], db, skip_existing: bool = False) -> Tuple[int, int, int, List[Dict]]:
    """Process a batch of expenses and return (imported_count, failed_count, uncategorized_count, failed_rows)"""
    imported_count = 0
    uncategorized_count = 0
    failed_rows = []
    
    for expense_data in expenses_batch:
        try:
            # Skip if amount is 0 or negative
            amount = expense_data.get('amount', 0)
            if amount <= 0:
                continue
            
            # Check for existing expense if skip_existing is enabled
            if skip_existing and 'id' in expense_data and expense_data.get('id'):
                try:
                    expense_id = UUID(str(expense_data['id']))
                    existing = db.query(Expense).filter(Expense.id == expense_id).first()
                    if existing:
                        logger.debug(f"Skipping existing expense ID: {expense_id}")
                        continue
                except (ValueError, TypeError):
                    pass
            
            # Match category
            matched_category = match_category(expense_data, all_categories, category_id_map, db)
            
            if matched_category:
                expense_data['category_id'] = matched_category.id
            else:
                expense_data['category_id'] = None
                uncategorized_count += 1
            
            # Ensure date is a date object (not datetime)
            expense_date = expense_data['date']
            if isinstance(expense_date, datetime):
                expense_date = expense_date.date()
            elif not isinstance(expense_date, date):
                expense_date = date.today()
            
            # Convert to ExpenseCreate format
            expense_create = ExpenseCreate(
                amount=expense_data['amount'],
                currency=expense_data.get('currency', 'IDR'),
                description=expense_data['description'],
                category_id=expense_data.get('category_id'),
                date=expense_date,
                tags=expense_data.get('tags', []),
                location=expense_data.get('location'),
                notes=expense_data.get('notes'),
                is_recurring=False
            )
            
            # Create expense
            db_expense = Expense(**expense_create.model_dump())
            db.add(db_expense)
            imported_count += 1
            
        except Exception as e:
            error_info = {
                "error": str(e),
                "data": expense_data
            }
            logger.exception(f"Failed to process expense: {str(e)}")
            failed_rows.append(error_info)
    
    # Commit the entire batch
    try:
        db.commit()
        logger.info(f"Batch committed: {imported_count} expenses imported")
    except Exception as e:
        db.rollback()
        logger.error(f"Batch commit failed: {str(e)}")
        # Mark all as failed if batch commit fails
        failed_rows.extend([{"error": f"Batch commit failed: {str(e)}", "data": exp} for exp in expenses_batch])
        imported_count = 0
        uncategorized_count = 0  # Reset if batch failed
    
    return imported_count, len(failed_rows), uncategorized_count, failed_rows


def print_progress(current: int, total: int, batch_num: int, total_batches: int):
    """Print progress information"""
    percentage = (current / total * 100) if total > 0 else 0
    print(f"\rProcessing batch {batch_num}/{total_batches} - Imported: {current}/{total} ({percentage:.1f}%)", end='', flush=True)


def get_database_url(args_db_url: Optional[str] = None) -> str:
    """Get database URL from arguments, environment variable, or prompt user"""
    # Check command-line argument first
    if args_db_url:
        return args_db_url
    
    # Check environment variable
    db_url = os.getenv("DATABASE_URL")
    if db_url:
        # Warn if it looks like a local SQLite database
        if db_url.startswith("sqlite"):
            logger.warning("⚠️  DATABASE_URL points to SQLite database. Make sure this is correct for production import!")
        return db_url
    
    # Prompt user
    print("\n⚠️  DATABASE_URL not found in environment variables.")
    print("Please provide your production database URL.")
    print("You can get it from Railway Dashboard → Backend Service → Variables tab")
    print("Or set it as: export DATABASE_URL='postgresql://...'")
    db_url = input("\nEnter DATABASE_URL: ").strip()
    
    if not db_url:
        raise ValueError("DATABASE_URL is required")
    
    return db_url


def main():
    parser = argparse.ArgumentParser(
        description="Import Excel file to production database",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Using environment variable
  export DATABASE_URL="postgresql://user:pass@host:port/dbname"
  poetry run python scripts/import_excel_local.py file.xlsx

  # With custom batch size
  poetry run python scripts/import_excel_local.py file.xlsx --batch-size 500

  # Skip existing expenses
  poetry run python scripts/import_excel_local.py file.xlsx --skip-existing

  # Custom database URL
  poetry run python scripts/import_excel_local.py file.xlsx --db-url "postgresql://..."
        """
    )
    
    parser.add_argument(
        "excel_file",
        type=str,
        help="Path to Excel file (.xlsx or .xls)"
    )
    
    parser.add_argument(
        "--batch-size",
        type=int,
        default=200,
        help="Number of expenses to process per batch (default: 200)"
    )
    
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Skip expenses that already exist (by ID)"
    )
    
    parser.add_argument(
        "--db-url",
        type=str,
        default=None,
        help="Database URL (overrides DATABASE_URL env var)"
    )
    
    parser.add_argument(
        "--log-file",
        type=str,
        default=None,
        help="Optional log file path for detailed logging"
    )
    
    args = parser.parse_args()
    
    # Setup file logging if requested
    if args.log_file:
        file_handler = logging.FileHandler(args.log_file)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(file_handler)
    
    # Validate Excel file
    excel_path = Path(args.excel_file)
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        sys.exit(1)
    
    if excel_path.suffix.lower() not in ['.xlsx', '.xls']:
        logger.error(f"Invalid file type. Expected .xlsx or .xls, got: {excel_path.suffix}")
        sys.exit(1)
    
    # Get database URL
    try:
        db_url = get_database_url(args.db_url)
    except ValueError as e:
        logger.error(str(e))
        sys.exit(1)
    
    # Check for internal Railway URL (won't work from local machine)
    if "postgres.railway.internal" in db_url or ".railway.internal" in db_url:
        logger.error("✗ ERROR: Internal Railway database URL detected!")
        logger.error("   The URL contains '.railway.internal' which only works within Railway's network.")
        logger.error("   You need the PUBLIC database URL for local access.")
        logger.error("")
        logger.error("   To get the public URL:")
        logger.error("   1. Go to Railway → PostgreSQL service → 'Connect' or 'Public Networking' tab")
        logger.error("   2. Copy the public connection string (hostname should NOT contain .railway.internal)")
        logger.error("   3. Or enable 'Public Networking' in Settings → Networking")
        logger.error("")
        sys.exit(1)
    
    # Security warning
    if "postgresql" in db_url.lower() and "@" in db_url:
        if "railway.app" in db_url or "railway" in db_url:
            logger.info("✓ Database URL detected (PostgreSQL - Railway)")
        else:
            logger.info("✓ Database URL detected (PostgreSQL)")
    else:
        logger.warning("⚠️  Database URL format looks unusual. Please verify it's correct.")
    
    # Override DATABASE_URL for database connection
    os.environ["DATABASE_URL"] = db_url
    
    # Recreate engine with production database URL
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    
    if db_url.startswith("sqlite"):
        engine = create_engine(db_url, connect_args={"check_same_thread": False}, echo=False)
    else:
        engine = create_engine(db_url, echo=False)
    
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    
    # Test database connection
    logger.info("Testing database connection...")
    try:
        from sqlalchemy import text
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
        logger.info("✓ Database connection successful")
    except Exception as e:
        logger.error(f"✗ Database connection failed: {str(e)}")
        logger.error("Please check your DATABASE_URL and ensure the database is accessible.")
        sys.exit(1)
    
    # Read Excel file
    logger.info(f"Reading Excel file: {excel_path}")
    try:
        with open(excel_path, 'rb') as f:
            file_content = f.read()
        logger.info(f"✓ File read successfully ({len(file_content) / 1024:.2f} KB)")
    except Exception as e:
        logger.error(f"✗ Failed to read Excel file: {str(e)}")
        sys.exit(1)
    
    # Load workbook to check for Categories sheet
    workbook = load_workbook(io.BytesIO(file_content), data_only=True)
    
    # Import categories if Categories sheet exists
    db = SessionLocal()
    category_id_map: Dict[str, UUID] = {}
    categories_imported = 0
    try:
        categories_imported = import_categories(workbook, db, category_id_map)
        logger.info(f"✓ Imported {categories_imported} categories")
    except Exception as e:
        logger.warning(f"Category import failed: {str(e)}")
        db.rollback()
    
    # Parse Excel file for expenses
    logger.info("Parsing Excel file for expenses...")
    import_service = ExcelImportService(file_content)
    expenses_data, parse_errors = import_service.parse()
    
    if parse_errors:
        logger.warning(f"Parse errors encountered: {len(parse_errors)}")
        for error in parse_errors[:5]:
            logger.warning(f"  - {error}")
    
    if not expenses_data:
        logger.error("No expenses found in Excel file")
        db.close()
        sys.exit(1)
    
    logger.info(f"✓ Found {len(expenses_data)} expense rows to import")
    
    # Load all categories for matching
    all_categories = db.query(Category).all()
    logger.info(f"✓ Loaded {len(all_categories)} categories from database")
    
    # Process expenses in batches
    batch_size = args.batch_size
    total_expenses = len(expenses_data)
    total_batches = (total_expenses + batch_size - 1) // batch_size
    
    logger.info(f"\nStarting batch import (batch size: {batch_size}, total batches: {total_batches})")
    logger.info("=" * 60)
    
    imported_count = 0
    failed_count = 0
    skipped_count = 0
    uncategorized_count = 0
    all_failed_rows = []
    category_matches: Dict[str, int] = {}
    
    for batch_num in range(total_batches):
        start_idx = batch_num * batch_size
        end_idx = min(start_idx + batch_size, total_expenses)
        batch = expenses_data[start_idx:end_idx]
        
        # Count skipped expenses (amount <= 0)
        skipped_in_batch = sum(1 for exp in batch if exp.get('amount', 0) <= 0)
        skipped_count += skipped_in_batch
        
        # Filter out skipped expenses
        batch = [exp for exp in batch if exp.get('amount', 0) > 0]
        
        if not batch:
            print_progress(imported_count, total_expenses, batch_num + 1, total_batches)
            continue
        
        # Process batch
        batch_imported, batch_failed, batch_uncategorized, batch_failed_rows = process_batch(
            batch, all_categories, category_id_map, db, args.skip_existing
        )
        
        imported_count += batch_imported
        failed_count += batch_failed
        uncategorized_count += batch_uncategorized
        all_failed_rows.extend(batch_failed_rows)
        
        # Track category matches
        for exp in batch:
            if exp.get('category_id'):
                cat = next((c for c in all_categories if c.id == exp.get('category_id')), None)
                if cat:
                    category_matches[cat.name] = category_matches.get(cat.name, 0) + 1
        
        # Print progress
        print_progress(imported_count, total_expenses, batch_num + 1, total_batches)
    
    print()  # New line after progress
    
    # Print summary
    logger.info("\n" + "=" * 60)
    logger.info("IMPORT SUMMARY")
    logger.info("=" * 60)
    logger.info(f"Total rows in file:     {len(expenses_data)}")
    logger.info(f"Successfully imported:  {imported_count}")
    logger.info(f"Failed:                 {failed_count}")
    logger.info(f"Skipped (amount <= 0):  {skipped_count}")
    logger.info(f"Uncategorized:         {uncategorized_count}")
    logger.info(f"Categories imported:    {categories_imported}")
    
    if category_matches:
        logger.info("\nCategory matches:")
        for cat_name, count in sorted(category_matches.items(), key=lambda x: x[1], reverse=True):
            logger.info(f"  - {cat_name}: {count}")
    
    if all_failed_rows:
        logger.warning(f"\nFailed rows ({len(all_failed_rows)}):")
        for failed_row in all_failed_rows[:10]:
            logger.warning(f"  - {failed_row.get('error', 'Unknown error')}")
        if len(all_failed_rows) > 10:
            logger.warning(f"  ... and {len(all_failed_rows) - 10} more")
    
    if parse_errors:
        logger.warning(f"\nParse errors ({len(parse_errors)}):")
        for error in parse_errors[:5]:
            logger.warning(f"  - {error}")
        if len(parse_errors) > 5:
            logger.warning(f"  ... and {len(parse_errors) - 5} more")
    
    db.close()
    
    if imported_count > 0:
        logger.info("\n✓ Import completed successfully!")
        sys.exit(0)
    else:
        logger.error("\n✗ No expenses were imported. Please check the errors above.")
        sys.exit(1)


if __name__ == "__main__":
    main()
