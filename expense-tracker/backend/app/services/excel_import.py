"""
Excel file import service
"""
from typing import List, Dict, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import re
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell
import io

logger = logging.getLogger(__name__)


class ExcelImportService:
    """Service to parse and extract expense data from Excel files"""
    
    # Common column name mappings (case-insensitive)
    COLUMN_MAPPINGS = {
        'date': ['date', 'tanggal', 'transaction date', 'expense date', 'timestamp'],
        'amount': ['amount', 'jumlah', 'total', 'price', 'harga', 'value'],
        'description': ['description', 'deskripsi', 'note', 'notes', 'detail', 'item', 'merchant', 'store'],
        'name': ['name'],  # Separate mapping for Name column
        'rawtext': ['rawtext', 'raw text'],  # Separate mapping for RawText column
        'category': ['category', 'kategori', 'type', 'jenis'],
        'payment_method': ['payment method', 'payment', 'method', 'cara bayar', 'metode pembayaran'],
        'location': ['location', 'lokasi', 'place', 'tempat'],
        'notes': ['notes', 'note', 'catatan', 'remarks', 'keterangan'],
        'who': ['who'],  # Separate mapping for Who column
        'currency': ['currency', 'mata uang', 'matauang'],
        'tags': ['tags', 'tag', 'label'],
    }
    
    def __init__(self, file_content: bytes):
        """
        Initialize with Excel file content
        
        Args:
            file_content: Binary content of Excel file
        """
        self.file_content = file_content
        self.workbook = None
        self.worksheet = None
        self.column_map: Dict[str, int] = {}
        self.errors: List[str] = []
    
    def parse(self) -> Tuple[List[Dict], List[str]]:
        """
        Parse Excel file and extract expense data
        
        Returns:
            Tuple of (expenses list, errors list)
        """
        logger.info("Starting Excel file parsing")
        logger.debug(f"File size: {len(self.file_content)} bytes")
        
        try:
            # Load workbook from bytes
            logger.debug("Loading workbook from bytes")
            self.workbook = load_workbook(io.BytesIO(self.file_content), data_only=True)
            logger.info(f"Workbook loaded successfully. Sheets: {self.workbook.sheetnames}")
            
            # Use first sheet
            self.worksheet = self.workbook.active
            logger.info(f"Using active sheet: {self.worksheet.title}, Max rows: {self.worksheet.max_row}, Max cols: {self.worksheet.max_column}")
            
            # Detect header row and map columns
            logger.debug("Detecting column headers")
            self._detect_columns()
            logger.info(f"Column mapping detected: {self.column_map}")
            
            if not self.column_map:
                error_msg = "Could not detect required columns. Please ensure your Excel file has columns for Date, Amount, and Description."
                logger.error(error_msg)
                return [], [error_msg]
            
            # Extract expenses
            expenses = []
            row_num = 2  # Start from row 2 (assuming row 1 is header)
            logger.info(f"Starting row extraction from row {row_num} to {self.worksheet.max_row}")
            
            while row_num <= self.worksheet.max_row:
                row_data = self._extract_row(row_num)
                
                if row_data:
                    logger.debug(f"Row {row_num}: Extracted data - {row_data}")
                    # Validate row data
                    validation_error = self._validate_row(row_data, row_num)
                    if validation_error:
                        error_msg = f"Row {row_num}: {validation_error}"
                        logger.warning(error_msg)
                        self.errors.append(error_msg)
                    else:
                        logger.debug(f"Row {row_num}: Validation passed, adding to expenses")
                        expenses.append(row_data)
                else:
                    logger.debug(f"Row {row_num}: Empty row, skipping")
                
                row_num += 1
            
            logger.info(f"Parsing complete. Extracted {len(expenses)} valid expenses, {len(self.errors)} errors")
            return expenses, self.errors
            
        except Exception as e:
            error_msg = f"Error parsing Excel file: {str(e)}"
            logger.exception(error_msg)
            return [], [error_msg]
    
    def _detect_columns(self):
        """Detect column headers and create mapping"""
        if not self.worksheet:
            logger.warning("No worksheet available for column detection")
            return
        
        logger.debug("Starting column detection")
        # Check first few rows for headers
        for row_num in range(1, min(4, self.worksheet.max_row + 1)):
            row = self.worksheet[row_num]
            logger.debug(f"Checking row {row_num} for headers")
            
            for col_idx, cell in enumerate(row, start=1):
                if not cell.value:
                    continue
                
                cell_value = str(cell.value).strip().lower()
                logger.debug(f"Row {row_num}, Column {col_idx}: '{cell_value}'")
                
                # Check against column mappings
                for field, aliases in self.COLUMN_MAPPINGS.items():
                    if cell_value in aliases and field not in self.column_map:
                        self.column_map[field] = col_idx
                        logger.info(f"Mapped column '{cell_value}' (col {col_idx}) to field '{field}'")
                        break
            
            # If we found required columns, stop searching
            # Check for description OR (name OR rawtext) as description source
            has_description = (
                'description' in self.column_map or 
                'name' in self.column_map or 
                'rawtext' in self.column_map
            )
            if 'date' in self.column_map and 'amount' in self.column_map and has_description:
                logger.info(f"Found all required columns in row {row_num}, stopping search")
                break
        
        if not self.column_map:
            logger.warning("No columns detected")
        else:
            logger.info(f"Column detection complete. Mapped {len(self.column_map)} columns")
    
    def _extract_row(self, row_num: int) -> Optional[Dict]:
        """Extract data from a single row"""
        logger.debug(f"Extracting row {row_num}")
        row = self.worksheet[row_num]
        row_data = {}
        
        # Extract each mapped field
        for field, col_idx in self.column_map.items():
            if col_idx <= len(row):
                cell = row[col_idx - 1]
                value = cell.value
                
                if value is not None:
                    row_data[field] = value
                    logger.debug(f"Row {row_num}: Extracted {field} = {value}")
        
        # Handle description field - prefer 'name' over 'rawtext' if both exist
        # Combine name and rawtext into description
        description_parts = []
        if 'name' in row_data and row_data['name']:
            name_value = str(row_data['name']).strip()
            description_parts.append(name_value)
            logger.debug(f"Row {row_num}: Found Name field: '{name_value}'")
        
        if 'rawtext' in row_data and row_data['rawtext']:
            rawtext_value = str(row_data['rawtext']).strip()
            logger.debug(f"Row {row_num}: Found RawText field: '{rawtext_value}'")
            # If rawtext contains amount, it's likely the full description
            # Otherwise, append it
            if rawtext_value and rawtext_value not in description_parts:
                if not description_parts:
                    description_parts.append(rawtext_value)
                    logger.debug(f"Row {row_num}: Using RawText as description (Name not available)")
                else:
                    # Store rawtext in notes if name exists
                    row_data['notes'] = rawtext_value
                    logger.debug(f"Row {row_num}: Storing RawText in notes (Name exists)")
        
        # Set description from available sources
        if description_parts:
            row_data['description'] = description_parts[0]
            logger.debug(f"Row {row_num}: Set description to '{row_data['description']}'")
        elif 'description' not in row_data or not row_data['description']:
            # Fallback to any available description field
            row_data['description'] = row_data.get('name') or row_data.get('rawtext') or ''
            logger.debug(f"Row {row_num}: Using fallback description: '{row_data['description']}'")
        
        # Handle 'who' field - add to notes if available
        if 'who' in row_data and row_data['who']:
            who_value = str(row_data['who']).strip()
            if who_value:
                existing_notes = row_data.get('notes', '')
                if existing_notes:
                    row_data['notes'] = f"{existing_notes} (by {who_value})"
                else:
                    row_data['notes'] = f"by {who_value}"
                logger.debug(f"Row {row_num}: Added Who field to notes: '{row_data['notes']}'")
        
        # Return None if row is empty
        if not row_data:
            logger.debug(f"Row {row_num}: No data extracted, returning None")
            return None
        
        logger.debug(f"Row {row_num}: Extraction complete. Fields: {list(row_data.keys())}")
        return row_data
    
    def _validate_row(self, row_data: Dict, row_num: int) -> Optional[str]:
        """Validate a row of expense data"""
        logger.debug(f"Row {row_num}: Starting validation")
        errors = []
        
        # Parse date - be lenient, try to infer from datetime
        if 'date' not in row_data or not row_data['date']:
            # Use today's date as fallback if no date provided
            from datetime import date as date_class
            row_data['date'] = date_class.today()
            logger.warning(f"Row {row_num}: No date field found, using today's date: {row_data['date']}")
        else:
            # Parse date - try to extract from datetime
            logger.debug(f"Row {row_num}: Parsing date: {row_data['date']}")
            parsed_date = self._parse_date(row_data['date'])
            if not parsed_date:
                # If parsing fails, try to extract date from string or use today
                from datetime import date as date_class
                # Try to extract just the date part from datetime string
                date_str = str(row_data['date']).strip()
                # Look for date pattern at the start (e.g., "1/1/2026" from "1/1/2026 20:55:47")
                date_match = re.match(r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})', date_str)
                if date_match:
                    date_part = date_match.group(1)
                    parsed_date = self._parse_date(date_part)
                
                if not parsed_date:
                    # Last resort: use today's date
                    parsed_date = date_class.today()
                    logger.warning(f"Row {row_num}: Could not parse date '{row_data['date']}', using today's date: {parsed_date}")
                else:
                    logger.info(f"Row {row_num}: Extracted date from datetime string: {parsed_date}")
            
            row_data['date'] = parsed_date
            logger.debug(f"Row {row_num}: Date set to: {parsed_date}")
        
        if 'amount' not in row_data or row_data['amount'] is None:
            errors.append("Amount is required")
            logger.warning(f"Row {row_num}: Missing amount field")
        else:
            # Parse amount
            logger.debug(f"Row {row_num}: Parsing amount: {row_data['amount']}")
            parsed_amount = self._parse_amount(row_data['amount'])
            if parsed_amount is None:
                errors.append(f"Invalid amount: {row_data['amount']}")
                logger.warning(f"Row {row_num}: Failed to parse amount: {row_data['amount']}")
            else:
                row_data['amount'] = parsed_amount
                # Allow 0 amounts (they will be skipped during import, but don't error here)
                logger.debug(f"Row {row_num}: Amount parsed successfully: {parsed_amount}")
        
        if 'description' not in row_data or not row_data['description']:
            errors.append("Description is required")
            logger.warning(f"Row {row_num}: Missing description field")
        else:
            # Normalize description
            row_data['description'] = str(row_data['description']).strip()
            if len(row_data['description']) > 500:
                errors.append("Description too long (max 500 characters)")
                logger.warning(f"Row {row_num}: Description too long: {len(row_data['description'])} chars")
            else:
                logger.debug(f"Row {row_num}: Description validated: '{row_data['description']}'")
        
        # Parse optional fields
        if 'payment_method' in row_data and row_data['payment_method']:
            row_data['payment_method'] = str(row_data['payment_method']).strip()
            logger.debug(f"Row {row_num}: Payment method: {row_data['payment_method']}")
        else:
            row_data['payment_method'] = 'Cash'  # Default
            logger.debug(f"Row {row_num}: Using default payment method: Cash")
        
        if 'currency' in row_data and row_data['currency']:
            currency = str(row_data['currency']).strip().upper()
            if len(currency) == 3:
                row_data['currency'] = currency
                logger.debug(f"Row {row_num}: Currency: {currency}")
            else:
                row_data['currency'] = 'IDR'  # Default
                logger.debug(f"Row {row_num}: Invalid currency format, using default: IDR")
        else:
            row_data['currency'] = 'IDR'  # Default
            logger.debug(f"Row {row_num}: No currency specified, using default: IDR")
        
        if 'location' in row_data and row_data['location']:
            location = str(row_data['location']).strip()
            if len(location) > 200:
                location = location[:200]
                logger.debug(f"Row {row_num}: Location truncated to 200 chars")
            row_data['location'] = location
        else:
            row_data['location'] = None
        
        if 'notes' in row_data and row_data['notes']:
            row_data['notes'] = str(row_data['notes']).strip()
            logger.debug(f"Row {row_num}: Notes: {row_data['notes'][:50]}...")
        else:
            row_data['notes'] = None
        
        if 'tags' in row_data and row_data['tags']:
            # Parse tags (comma-separated or space-separated)
            tags_str = str(row_data['tags']).strip()
            if ',' in tags_str:
                tags = [t.strip() for t in tags_str.split(',')]
            else:
                tags = tags_str.split()
            row_data['tags'] = [t for t in tags if t]
            logger.debug(f"Row {row_num}: Parsed tags: {row_data['tags']}")
        else:
            row_data['tags'] = []
        
        if errors:
            error_msg = "; ".join(errors)
            logger.warning(f"Row {row_num}: Validation failed - {error_msg}")
            return error_msg
        
        logger.debug(f"Row {row_num}: Validation passed")
        return None
    
    def _parse_date(self, value) -> Optional[date]:
        """Parse date from various formats - lenient parsing that extracts date from datetime"""
        logger.debug(f"Parsing date value: {value} (type: {type(value)})")
        
        if isinstance(value, date):
            logger.debug(f"Value is already a date: {value}")
            return value
        if isinstance(value, datetime):
            logger.debug(f"Value is datetime, extracting date: {value.date()}")
            return value.date()
        
        if not value:
            logger.debug("Date value is empty")
            return None
        
        value_str = str(value).strip()
        logger.debug(f"Date string to parse: '{value_str}'")
        
        # Try parsing datetime with time component first (e.g., "1/1/2026 20:55:47")
        # This is the most common format from Excel Timestamp columns
        datetime_formats = [
            '%m/%d/%Y %H:%M:%S',      # 1/1/2026 20:55:47
            '%d/%m/%Y %H:%M:%S',      # 1/1/2026 20:55:47 (DD/MM/YYYY)
            '%Y-%m-%d %H:%M:%S',      # 2026-01-01 20:55:47
            '%m/%d/%Y %H:%M',         # 1/1/2026 20:55
            '%d/%m/%Y %H:%M',         # 1/1/2026 20:55 (DD/MM/YYYY)
            '%Y-%m-%d %H:%M',         # 2026-01-01 20:55
            '%m-%d-%Y %H:%M:%S',      # 1-1-2026 20:55:47
            '%d-%m-%Y %H:%M:%S',      # 1-1-2026 20:55:47 (DD/MM/YYYY)
        ]
        
        for fmt in datetime_formats:
            try:
                parsed = datetime.strptime(value_str, fmt).date()
                logger.debug(f"Successfully parsed date with datetime format '{fmt}': {parsed}")
                return parsed
            except ValueError:
                continue
        
        # Try common date formats (without time)
        date_formats = [
            '%Y-%m-%d',      # 2026-01-01
            '%m/%d/%Y',      # 1/1/2026
            '%d/%m/%Y',      # 1/1/2026 (DD/MM/YYYY)
            '%d-%m-%Y',      # 1-1-2026
            '%Y/%m/%d',      # 2026/01/01
            '%d.%m.%Y',      # 1.1.2026
            '%m-%d-%Y',      # 1-1-2026
        ]
        
        for fmt in date_formats:
            try:
                parsed = datetime.strptime(value_str, fmt).date()
                logger.debug(f"Successfully parsed date with format '{fmt}': {parsed}")
                return parsed
            except ValueError:
                continue
        
        # Try parsing as Excel date serial number
        try:
            serial = float(value_str)
            logger.debug(f"Trying to parse as Excel serial number: {serial}")
            # Excel epoch is 1900-01-01, but Excel incorrectly treats 1900 as a leap year
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            parsed = (excel_epoch + timedelta(days=int(serial))).date()
            logger.debug(f"Successfully parsed Excel serial number: {parsed}")
            return parsed
        except (ValueError, OverflowError) as e:
            logger.debug(f"Failed to parse as Excel serial number: {e}")
        
        # Try to extract date from string using regex (e.g., extract "1/1/2026" from "1/1/2026 20:55:47")
        date_patterns = [
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})',  # 1/1/2026 or 1-1-2026
            r'(\d{4}[/-]\d{1,2}[/-]\d{1,2})',  # 2026/1/1 or 2026-1-1
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, value_str)
            if match:
                date_part = match.group(1)
                logger.debug(f"Extracted date part from string: '{date_part}'")
                # Try parsing the extracted date part
                for fmt in date_formats:
                    try:
                        parsed = datetime.strptime(date_part, fmt).date()
                        logger.debug(f"Successfully parsed extracted date part with format '{fmt}': {parsed}")
                        return parsed
                    except ValueError:
                        continue
        
        logger.warning(f"Could not parse date from value: {value_str}")
        return None
    
    def _parse_amount(self, value) -> Optional[float]:
        """Parse amount from various formats"""
        logger.debug(f"Parsing amount value: {value} (type: {type(value)})")
        
        if isinstance(value, (int, float)):
            parsed = float(value)
            logger.debug(f"Value is already numeric: {parsed}")
            return parsed
        
        if not value:
            logger.debug("Amount value is empty")
            return None
        
        value_str = str(value).strip()
        logger.debug(f"Amount string to parse: '{value_str}'")
        
        # Remove currency symbols and spaces
        original_str = value_str
        value_str = re.sub(r'[Rp$€£¥,\s]', '', value_str)
        if original_str != value_str:
            logger.debug(f"Removed currency symbols: '{value_str}'")
        
        # Replace period with nothing if it's a thousands separator (e.g., 1.000.000)
        # Check if there are multiple periods
        if value_str.count('.') > 1:
            value_str = value_str.replace('.', '')
            logger.debug(f"Removed multiple periods (thousands separator): '{value_str}'")
        # If single period, check if it's likely a decimal separator
        elif '.' in value_str:
            parts = value_str.split('.')
            # If part after period has more than 2 digits, it's likely thousands separator
            if len(parts) > 1 and len(parts[1]) > 2:
                value_str = value_str.replace('.', '')
                logger.debug(f"Removed period (thousands separator): '{value_str}'")
        
        try:
            parsed = float(value_str)
            logger.debug(f"Successfully parsed amount: {parsed}")
            return parsed
        except ValueError as e:
            logger.warning(f"Failed to parse amount from '{value_str}': {e}")
            return None
